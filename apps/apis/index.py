from flask import Blueprint, render_template, current_app
import pandas as pd
from pyecharts import options as opts
from pyecharts.charts import Bar,Tab
from pyecharts.components import Table
from openpyxl import load_workbook

index_bp = Blueprint('index', __name__, url_prefix='/')

def read_data(progress_name):
    file_path = current_app.config["DATA_FILE_PATH"] # 获取数据文件路径
    data = pd.read_excel(io=file_path, sheet_name=progress_name) # 读取指定流程的数据
    return data


def read_tech_details():
    data_details = read_data('技术详情')
    tech_name_details = data_details['技术名称'].tolist()
    tech_details = data_details['技术详情'].tolist()
    tech_details = list(zip(tech_name_details, tech_details))
    return tech_details

# view 前后端耦合较高
@index_bp.route('/')
def index():
    return render_template('index.html')

@index_bp.route('/dashboard')
def dashboard():
    file_path = current_app.config["DATA_FILE_PATH"] # 获取数据文件路径
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames[:-3] # 排除后三个拿来做预测的表单
    tech_nums = []

    progress_names = sheet_names[2:]
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        num_rows = sheet.max_row - 1
        tech_nums.append(num_rows)
    tech_nums = tech_nums[2:]
    bar = (
        Bar(init_opts=opts.InitOpts(width="100%"))
        .add_xaxis(progress_names)
        .add_yaxis("技术数量", tech_nums)
        .reversal_axis()
        .set_series_opts(label_opts=opts.LabelOpts(position="right"))
        .set_global_opts(title_opts=opts.TitleOpts(title="各流程技术数量"))
    )

    progress_tech = []
    tech_names = pd.read_excel(io=file_path, sheet_name="总数据")["技术名称"].tolist()
    final_table = []
    for sheet_name in progress_names:
        data = pd.read_excel(io=file_path, sheet_name=sheet_name)
        data = data["技术名称"].tolist()
        progress_tech.append(data)
    for tech_name in tech_names:
        final_list = [tech_name]
        for each_progress_tech in progress_tech:
            if tech_name in each_progress_tech:
                final_list.append("√")
            else:
                final_list.append("×")
        final_table.append(final_list)

    tech_details = read_tech_details()
    return render_template('dashboard.html', mybar=bar.render_embed()[:-2], final_table=final_table,
                           tech_details=tech_details)